import logging
from openpyxl import Workbook, load_workbook


logging.getLogger(__name__)


def save_xl(filename, xldata, captions=None):
    """Save list of lists to excel file

    Args:
        filename (string): File name or /path/to/file
        xldata ([[],[], ..., []]): Rows to save
        captions (list, optional): If first row is a captions. \
            Defaults to None.

    Returns:
        string: File name or path
    """
    try:
        if xldata != []:
            wb = Workbook()
            ws = wb.active
            if captions is not None:
                ws.append(captions)
            for data_row in xldata:
                ws.append(data_row)
            wb.save(filename)
            return filename
        else:
            logging.warning('Файл не сохранён, список строк пустой!')
    except Exception as e:
        logging.critical(f'Не могу сохранить файл: {filename}')
        logging.critical(e)


def read_xl(filename, is_captions=True):
    """Read excel file

    Args:
        filename (string): File name or /path/to/file
        is_captions (bool, optional): If first row is captions. \
            Defaults to True.

    Returns:
        dict: {
            "captions": [],
            "rows": [[],[], ..., []]
        }
    """
    try:
        rows = None
        wb = load_workbook(filename=filename, read_only=True)
        ws = wb.active
        ws_rows = ws.rows
        if is_captions:
            next(ws_rows)
        rows = [(row[0].value, row[1].value) for row in ws_rows]
        logging.info(rows)
        return rows
    except IOError as e:
        logging.critical(f'Не могу прочитать файл: {filename}')
        logging.critical(e)
    except Exception as e:
        logging.critical(f'Что-то пошло не так при чтении файла: {filename}')
        logging.critical(e)
    return []
